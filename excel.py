from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime, date


    
def to_excel(trade_id, entry_price, side, qty, start, end, pnl):
    wb = load_workbook(r"C:\Users\admin\Desktop\Trades\TradeHistory.xlsx")
    ws = wb.active
    row1 = ["Trade ID", "Entry Price", "Quantity", "Side", "Start", "End", "PNL"]

    # Create a new sheet if it's the new day
    if ws.title != str(date.today()):
        ws = wb.create_sheet(f"{date.today()}")
        ws.append(row1)

    # Bold first row
    asc = 65
    for i in range(1, len(row1)+1):
        symbol = chr(asc)
        cell = symbol+"1"
        ws[cell].font = Font(bold=True)
        asc += 1
    
    # Append trade information into cells
    ws.append([trade_id, entry_price, qty, side, start, end, pnl])
    wb.save(r"C:\Users\admin\Desktop\Trades\TradeHistory.xlsx")

